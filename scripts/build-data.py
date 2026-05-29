#!/usr/bin/env python3
"""
build-data.py — Pivot v2 data pipeline

Builds the JSON files under docs/public/data/v2/ from two public sources:
  1. Eloundou, Manning, Mishkin, Rock (2023) "GPTs are GPTs" — task-level GPT
     exposure labels, O*NET task statements, BLS national wages (May 2021),
     BLS Employment Projections. https://github.com/openai/GPTs-are-GPTs
  2. Felten, Raj, Seamans (2021) AIOE scores by occupation.
     https://github.com/AIOE-Data/AIOE

Usage:
    python3 scripts/build-data.py

Run once after pulling. Source data is downloaded into ./data-sources/
(gitignored) on first run.

Selection: top ~200 BLS detailed occupations by total employment.

Outputs (under docs/public/data/v2/):
  meta.json          - build timestamp, source versions, license attribution
  occupations.json   - master list: SOC, title, AIOE, GPT exposure α/β/γ
  tasks.json         - per-occupation task array with GPT-4 exposure label
  wages.json         - per-occupation annual P10/P25/P50/P75/P90 (national)
  projections.json   - per-occupation 2020-30 employment % change
"""

import csv
import io
import json
import os
import sys
import urllib.request
import zipfile
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
SOURCES = ROOT / "data-sources"
OUT = ROOT / "docs" / "public" / "data" / "v2"
SOURCES.mkdir(exist_ok=True)
OUT.mkdir(parents=True, exist_ok=True)

ELOUNDOU_ZIP = SOURCES / "eloundou-gpts-are-gpts.zip"
AIOE_XLSX = SOURCES / "aioe-data-appendix.xlsx"
ELOUNDOU_URL = "https://codeload.github.com/openai/GPTs-are-GPTs/zip/refs/heads/main"
AIOE_URL = "https://raw.githubusercontent.com/AIOE-Data/AIOE/main/AIOE_DataAppendix.xlsx"

TOP_N = 200  # how many detailed occupations to include in v1


def download(url: str, dest: Path) -> None:
    if dest.exists() and dest.stat().st_size > 1000:
        print(f"  cached: {dest.name}")
        return
    print(f"  fetching: {url}")
    req = urllib.request.Request(url, headers={"User-Agent": "1day1doom-build/1"})
    with urllib.request.urlopen(req, timeout=60) as r, open(dest, "wb") as f:
        f.write(r.read())


def open_eloundou(name: str):
    """Open a file inside the Eloundou zip."""
    with zipfile.ZipFile(ELOUNDOU_ZIP) as z:
        # Files live under GPTs-are-GPTs-main/
        for info in z.namelist():
            if info.endswith("/" + name):
                return z.read(info).decode("utf-8", errors="replace")
    raise FileNotFoundError(name)


def load_aioe():
    """SOC 6-digit → AIOE score. Returns dict[soc_short] = aioe."""
    from openpyxl import load_workbook
    wb = load_workbook(AIOE_XLSX, read_only=True, data_only=True)
    ws = wb["Appendix A"]
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        soc, title, score = row[0], row[1], row[2]
        if soc and score is not None:
            out[str(soc).strip()] = float(score)
    return out


def load_occ_level():
    """O*NET-SOC → GPT exposure α/β/γ + title."""
    text = open_eloundou("occ_level.csv")
    out = {}
    for row in csv.DictReader(io.StringIO(text)):
        code = row["O*NET-SOC Code"]
        out[code] = {
            "title": row["Title"],
            "exp_alpha": float(row["dv_rating_alpha"]),
            "exp_beta": float(row["dv_rating_beta"]),
            "exp_gamma": float(row["dv_rating_gamma"]),
        }
    return out


def load_wages():
    """OCC_CODE (SOC 6-digit) → wages dict. Uses national row only."""
    text = open_eloundou("national_May2021_dl.csv")
    out = {}
    # The first column starts with BOM; csv reader handles it via utf-8-sig if we strip.
    text = text.lstrip("﻿")
    for row in csv.DictReader(io.StringIO(text)):
        if row.get("AREA_TYPE") != "1":
            continue  # 1 = national
        if row.get("O_GROUP") != "detailed":
            continue
        if row.get("I_GROUP") != "cross-industry":
            continue
        code = row.get("OCC_CODE", "").strip()
        if not code or code == "00-0000":
            continue
        def num(k):
            v = row.get(k, "").replace(",", "").strip()
            if not v or v in ("*", "#", "**"):
                return None
            try:
                return int(v)
            except ValueError:
                return None
        out[code] = {
            "title": row.get("OCC_TITLE", ""),
            "tot_emp": num("TOT_EMP"),
            "annual_p10": num("A_PCT10"),
            "annual_p25": num("A_PCT25"),
            "annual_median": num("A_MEDIAN"),
            "annual_p75": num("A_PCT75"),
            "annual_p90": num("A_PCT90"),
        }
    return out


def load_projections():
    """Occupation title (BLS-formatted, normalized to lowercase) → 2020-30 % change."""
    text = open_eloundou("occupations_projections_processed.csv")
    out = {}
    for row in csv.DictReader(io.StringIO(text)):
        name = (row.get("occupation") or "").strip().lower()
        pct = (row.get("pct_emp_change_2020_2030") or "").strip()
        if not name or not pct:
            continue
        try:
            out[name] = float(pct)
        except ValueError:
            continue
    return out


def load_tasks():
    """O*NET-SOC → list of tasks. Each task has GPT-4 + human automation label."""
    text = open_eloundou("automation_gpt4_human_labels.tsv")
    out = {}
    rdr = csv.DictReader(io.StringIO(text), delimiter="\t")
    for row in rdr:
        code = row["O*NET-SOC Code"]
        out.setdefault(code, []).append({
            "task_id": row["Task ID"],
            "task": row["Task"],
            "type": row["Task Type"],
            "gpt4": row["gpt4_automation"],   # T0/T1/T2/T3/T4
            "human": row["human_automation"],
        })
    return out


def soc_short(onet_soc: str) -> str:
    """O*NET-SOC '11-1011.00' -> '11-1011'."""
    return onet_soc.split(".")[0]


def build():
    print("step 1/4 downloading sources")
    download(ELOUNDOU_URL, ELOUNDOU_ZIP)
    download(AIOE_URL, AIOE_XLSX)

    print("step 2/4 parsing sources")
    aioe = load_aioe()
    occs = load_occ_level()
    wages = load_wages()
    projs = load_projections()
    tasks = load_tasks()
    print(f"  AIOE: {len(aioe)} | occ_level: {len(occs)} | wages: {len(wages)} | "
          f"projections: {len(projs)} | tasks (occupations): {len(tasks)}")

    print(f"step 3/4 selecting top {TOP_N} occupations by employment")
    # Pick top N by BLS total employment. Filter to occupations that exist in
    # all required tables (occ_level + wages + tasks).
    candidates = []
    for soc, w in wages.items():
        if w["tot_emp"] is None:
            continue
        # find matching O*NET-SOC code (soc is short 6-digit; occs keys end with .00 etc.)
        onet_match = next((k for k in occs.keys() if k.startswith(soc + ".")), None)
        if not onet_match:
            continue
        if onet_match not in tasks:
            continue
        candidates.append((w["tot_emp"], soc, onet_match))
    candidates.sort(reverse=True)
    selected = candidates[:TOP_N]
    print(f"  selected: {len(selected)} occupations")

    occupations = []
    out_tasks = {}
    out_wages = {}
    out_projs = {}
    for tot_emp, soc, onet_soc in selected:
        meta = occs[onet_soc]
        w = wages[soc]
        occupations.append({
            "soc": soc,
            "onet_soc": onet_soc,
            "title": meta["title"],
            "tot_emp": tot_emp,
            "aioe": round(aioe.get(soc), 4) if soc in aioe else None,
            "exp_alpha": round(meta["exp_alpha"], 4),
            "exp_beta": round(meta["exp_beta"], 4),
            "exp_gamma": round(meta["exp_gamma"], 4),
        })
        # Tasks: keep statement, type, gpt4 label only — drop human label
        # (kept in source for transparency, not needed in v1 UI).
        out_tasks[soc] = [
            {
                "id": t["task_id"],
                "task": t["task"],
                "type": t["type"],
                "gpt4": t["gpt4"],
            }
            for t in tasks[onet_soc]
        ]
        out_wages[soc] = {
            "p10": w["annual_p10"],
            "p25": w["annual_p25"],
            "median": w["annual_median"],
            "p75": w["annual_p75"],
            "p90": w["annual_p90"],
        }
        # Projection lookup by BLS title — projections dict is lowercase-normalized.
        # Try O*NET title first, then BLS OCC_TITLE fallback. Either may match
        # because the BLS Projections file uses lowercase plural forms.
        t1 = meta["title"].lower()
        t2 = w["title"].lower()
        out_projs[soc] = projs.get(t1) or projs.get(t2)

    print("step 4/4 writing JSON")
    meta = {
        "build_time": datetime.utcnow().isoformat() + "Z",
        "version": "v2.0.0-data",
        "selection": f"top {TOP_N} BLS detailed occupations by total employment",
        "sources": [
            {
                "id": "eloundou-2023",
                "name": "GPTs are GPTs: An Early Look at the Labor Market Impact Potential of Large Language Models",
                "authors": "Eloundou T, Manning S, Mishkin P, Rock D (2023)",
                "url": "https://github.com/openai/GPTs-are-GPTs",
                "license": "MIT",
                "provides": "O*NET task statements, GPT-4 task exposure labels, BLS national wages (May 2021), BLS Employment Projections 2020-2030",
            },
            {
                "id": "felten-2021-aioe",
                "name": "AIOE: Occupational, Industry, and Geographic Exposure to AI",
                "authors": "Felten E, Raj M, Seamans R (2021)",
                "url": "https://github.com/AIOE-Data/AIOE",
                "license": "Citation-requested (academic data)",
                "provides": "AIOE score per 6-digit SOC occupation",
            },
        ],
        "exposure_label_legend": {
            "T0": "No exposure — LLMs and tools that build on LLMs offer no or minimal improvement to time required.",
            "T1": "Direct exposure — LLM alone could reduce time to do the task by ≥50%.",
            "T2": "LLM+ exposed — LLM with tools / interfaces could reduce time by ≥50%.",
            "T3": "Image+ exposed (rare in occ-level set).",
            "T4": "All other partial-exposure or uncertain.",
        },
    }
    (OUT / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2))
    (OUT / "occupations.json").write_text(json.dumps(occupations, ensure_ascii=False, indent=2))
    (OUT / "tasks.json").write_text(json.dumps(out_tasks, ensure_ascii=False, indent=2))
    (OUT / "wages.json").write_text(json.dumps(out_wages, ensure_ascii=False, indent=2))
    (OUT / "projections.json").write_text(json.dumps(out_projs, ensure_ascii=False, indent=2))

    for name in ("meta", "occupations", "tasks", "wages", "projections"):
        f = OUT / f"{name}.json"
        print(f"  {name}.json: {f.stat().st_size / 1024:.1f} KB")
    print("done.")


if __name__ == "__main__":
    try:
        build()
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
